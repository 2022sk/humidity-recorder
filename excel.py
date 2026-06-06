"""엑셀 생성 – DB 레코드 기반"""

import io, os, math
from datetime import date, timedelta
from PIL import Image

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker

DAYS_KO = ["월","화","수","목","금","토","일"]
SLOTS   = ["오전1","오전2","오후1","오후2"]


def _excel_font() -> str:
    # 환경변수로 재정의 가능: EXCEL_FONT=맑은 고딕
    env = os.environ.get("EXCEL_FONT", "")
    return env if env else "현대하모니 Head"


def _heat_index(Ta: float, RH: float) -> float:
    Tw = (Ta * math.atan(0.151977*(RH+8.313659)**0.5)
          + math.atan(Ta+RH) - math.atan(RH-1.67633)
          + 0.00391838*RH**1.5*math.atan(0.023101*RH) - 4.686035)
    return round(-0.2442+0.55399*Tw+0.45535*Ta-0.0022*Tw**2+0.00278*Tw*Ta+3.0, 1)


def _heat_level(fl: float) -> str:
    if fl >= 38: return "위험"
    if fl >= 35: return "경고"
    if fl >= 33: return "주의"
    if fl >= 31: return "관심"
    return "-"


def week_label_ko(n: int) -> str:
    return ["첫째","둘째","셋째","넷째","다섯째"][min(n-1,4)] + "주"


def get_week_n(monday: date) -> int:
    """해당 월에서 몇 번째 주인지"""
    import calendar
    first = date(monday.year, monday.month, 1)
    mon   = first - timedelta(days=first.weekday())
    n = 1
    while mon < monday:
        mon += timedelta(days=7)
        n   += 1
    return n


def build_excel(records: list, meta: dict, monday: date) -> bytes:
    """records: DB에서 가져온 dict 리스트. 각 record에 _bytes(사진) 포함 가능."""
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "체감온도 기록관리 대장"
    FM = _excel_font()

    C_INK="1D1D1F"; C_BLUE="0071E3"; C_DARK="1D1D1F"; C_WHITE="FFFFFF"
    C_GRAY6="F5F5F7"; C_ROW0="FFFFFF"; C_ROW1="F9F9FB"; C_DATE="EBF3FF"; C_BORDER="C7C7CC"

    thin = Side(style="thin",color=C_BORDER)
    bdr  = Border(left=thin,right=thin,top=thin,bottom=thin)
    ctr  = Alignment(horizontal="center",vertical="center",wrap_text=False)
    lft  = Alignment(horizontal="left",  vertical="center",wrap_text=False,shrink_to_fit=True)
    SZ=10

    f_title = Font(bold=True,color=C_WHITE,name=FM,size=13)
    f_head  = Font(bold=True,color=C_WHITE,name=FM,size=SZ)
    f_meta  = Font(bold=True,color=C_INK,  name=FM,size=SZ)
    f_date  = Font(bold=True,color=C_BLUE, name=FM,size=SZ)
    f_body  = Font(         color=C_INK,  name=FM,size=SZ)
    f_sig   = Font(         color=C_INK,  name=FM,size=SZ)
    f_slot  = Font(bold=True,color=C_INK,  name=FM,size=SZ)  # 슬롯 헤더: 다크 텍스트

    fill_title    = PatternFill("solid",fgColor=C_DARK)
    fill_head     = PatternFill("solid",fgColor=C_BLUE)
    fill_photo_hdr= PatternFill("solid",fgColor="5B7DB8")   # 사진대지: 차분한 청색
    fill_photo_cell=PatternFill("solid",fgColor=C_WHITE)     # 사진 셀: 흰색
    fill_slots    = [                                         # 파스텔 슬롯 색
        PatternFill("solid",fgColor="BEE0F4"),  # 오전1: 연하늘
        PatternFill("solid",fgColor="BBE5D0"),  # 오전2: 연민트
        PatternFill("solid",fgColor="FAD9A0"),  # 오후1: 연살구
        PatternFill("solid",fgColor="DEC8F0"),  # 오후2: 연라벤더
    ]
    fill_meta  = PatternFill("solid",fgColor=C_GRAY6)
    fill_sig   = PatternFill("solid",fgColor=C_GRAY6)
    fill_date  = PatternFill("solid",fgColor=C_DATE)
    fill_date1 = PatternFill("solid",fgColor="E3EDFF")
    lv_colors  = {"위험":("FF3B30",C_WHITE),"경고":("FF9500",C_WHITE),
                  "주의":("FFCC00",C_INK), "관심":("34C759",C_WHITE)}

    ROW_H=34; NCOL=11; PC=NCOL+1; PCW=22
    HDR=["작성일","구분","측정시각","온도(°C)","습도(%)","체감온도(°C)","단계","조치사항","기타내용","측정자","비고"]
    CWIDTHS=[13,10,12,7.5,7.5,10,8,17,13,10,13]
    for i,w in enumerate(CWIDTHS,1): ws.column_dimensions[get_column_letter(i)].width=w
    for i in range(4): ws.column_dimensions[get_column_letter(PC+i)].width=PCW

    wk_n = get_week_n(monday)
    wk_lbl = f"{monday.year}년 {monday.month}월 {week_label_ko(wk_n)}"

    # 행1: 제목
    ws.merge_cells(f"A1:{get_column_letter(PC+3)}1")
    c=ws["A1"]; c.value="체감온도 기록관리 대장"
    c.font=f_title; c.alignment=ctr; c.fill=fill_title; c.border=bdr
    ws.row_dimensions[1].height=38

    # 행2: 메타
    ws.merge_cells("A2:C2"); ws.merge_cells("D2:F2")
    ws.merge_cells("G2:I2"); ws.merge_cells("J2:K2")
    for addr,val in [("A2",f"현장명: {meta.get('현장명','')}"),
                     ("D2",f"업체명: {meta.get('업체명','')}"),
                     ("G2",f"측정위치: {meta.get('위치','')}"),
                     ("J2",wk_lbl)]:
        c=ws[addr]; c.value=val; c.font=f_meta; c.alignment=lft; c.border=bdr; c.fill=fill_meta
    ws.merge_cells(f"{get_column_letter(PC)}2:{get_column_letter(PC+3)}2")
    ph=ws[f"{get_column_letter(PC)}2"]
    ph.value="사진대지"; ph.fill=fill_photo_hdr; ph.font=f_head; ph.alignment=ctr; ph.border=bdr
    ws.row_dimensions[2].height=22

    # 행3: 헤더
    for ci,h in enumerate(HDR,1):
        c=ws.cell(row=3,column=ci,value=h)
        c.fill=fill_head; c.font=f_head; c.alignment=ctr; c.border=bdr
    for i,lbl in enumerate(SLOTS):
        c=ws.cell(row=3,column=PC+i,value=lbl)
        c.fill=fill_slots[i]; c.font=f_slot; c.alignment=ctr; c.border=bdr
    ws.row_dimensions[3].height=22

    dv=DataValidation(type="list",
        formula1='"N/A,추가휴식시간부여,보냉장구지급,작업시간대조정,작업중지,기타"',
        allow_blank=True,showDropDown=False)
    ws.add_data_validation(dv)

    # record map: (date_str, slot) → record
    rec_map: dict[tuple,dict] = {}
    for r in records:
        key = (r.get("measure_date","") or r.get("_date",""),
               r.get("slot","")         or r.get("_slot",""))
        rec_map[key] = r

    row_num=4
    for day_i in range(7):
        d=monday+timedelta(days=day_i); rs=row_num; re_=row_num+3
        even=(day_i%2==0)
        dfill=PatternFill("solid",fgColor=C_ROW0 if even else C_ROW1)

        empty_fill = PatternFill("solid", fgColor="E8E8E8")

        for slot in SLOTS:
            rec=rec_map.get((d.isoformat(),slot),{})
            act=rec.get("action","") or rec.get("조치사항","")
            T  =rec.get("temperature") if rec.get("temperature") is not None else rec.get("온도(°C)")
            RH =rec.get("humidity")    if rec.get("humidity")    is not None else rec.get("습도(%)")
            fl = rec.get("feels_like")
            hl = rec.get("heat_level","") or ""
            has_data = T is not None or RH is not None
            row_fill = dfill if has_data else empty_fill

            row_data=[f"{d.month}/{d.day}({DAYS_KO[d.weekday()]})",slot,
                      rec.get("measure_time","") or rec.get("측정시각",""),
                      T if T is not None else "", RH if RH is not None else "",
                      "","", act if act else "N/A",
                      rec.get("other_content","") or rec.get("기타내용","") if act=="기타" else "",
                      rec.get("measurer","") or rec.get("측정자",""),
                      rec.get("notes","")    or rec.get("비고","")]

            for ci,val in enumerate(row_data,1):
                c=ws.cell(row=row_num,column=ci,value=val)
                c.font=f_body; c.alignment=ctr; c.border=bdr
                c.fill = dfill if ci == 1 else row_fill  # A열(날짜)은 항상 day fill
                if ci in (4,5) and val!="": c.number_format="0.0"

            # 체감온도: DB 저장값 우선, 없으면 T·RH로 직접 계산
            if fl is None and T is not None and RH is not None:
                try: fl = _heat_index(float(T), float(RH))
                except Exception: fl = None
            # 단계: DB 저장값 우선, 없으면 체감온도로 계산 (수식 미사용)
            if hl not in ("위험","경고","주의","관심"):
                hl = _heat_level(fl) if fl is not None else ""

            fc=ws.cell(row=row_num,column=6)
            fc.value = fl if fl is not None else ""
            fc.font=f_body; fc.alignment=ctr; fc.border=bdr; fc.fill=row_fill
            if fl is not None: fc.number_format="0.0"

            gc=ws.cell(row=row_num,column=7)
            gc.value = hl
            gc.font=f_body; gc.alignment=ctr; gc.border=bdr; gc.fill=row_fill
            dv.add(ws.cell(row=row_num,column=8))
            row_num+=1

        ws.merge_cells(start_row=rs,start_column=1,end_row=re_,end_column=1)
        dc=ws.cell(row=rs,column=1)
        dc.font=f_date; dc.alignment=ctr
        dc.fill=fill_date if even else fill_date1; dc.border=bdr

        for si,slot in enumerate(SLOTS):
            col=PC+si
            ws.merge_cells(start_row=rs,start_column=col,end_row=re_,end_column=col)
            tc=ws.cell(row=rs,column=col); tc.border=bdr; tc.alignment=ctr; tc.fill=fill_photo_cell
            for ir in range(rs+1, re_+1):
                ic=ws.cell(row=ir,column=col)
                ic.border=bdr           # L~O 모두 동일한 테두리로 통일
                ic.fill=fill_photo_cell
            rec=rec_map.get((d.isoformat(),slot))
            if rec and rec.get("_bytes"):
                try:
                    pil=Image.open(io.BytesIO(rec["_bytes"])).convert("RGB")
                    buf=io.BytesIO(); pil.save(buf,format="PNG"); buf.seek(0)
                    xi=XLImage(buf); anc=TwoCellAnchor()
                    anc._from=AnchorMarker(col=col-1,colOff=0,row=rs-1,rowOff=0)
                    anc.to   =AnchorMarker(col=col,  colOff=0,row=re_,  rowOff=0)
                    anc.editAs="twoCell"; xi.anchor=anc; ws.add_image(xi)
                except Exception: tc.value="⚠"

    for r in range(4,row_num): ws.row_dimensions[r].height=ROW_H

    g_range=f"G4:G{row_num-1}"
    for lv,(bg,fg) in lv_colors.items():
        ws.conditional_formatting.add(g_range,CellIsRule(
            operator="equal",formula=[f'"{lv}"'],
            fill=PatternFill("solid",fgColor=bg),
            font=Font(color=fg,bold=True,name=FM,size=SZ)))

    leg=row_num
    for lbl,c1,c2,bg,fg in [
        ("KOSHA 폭염 단계",1,2,C_DARK,C_WHITE),("■ 관심  31°C~",3,4,"34C759",C_WHITE),
        ("■ 주의  33°C~",5,6,"FFCC00",C_INK),  ("■ 경고  35°C~",7,8,"FF9500",C_WHITE),
        ("■ 위험  38°C~",9,11,"FF3B30",C_WHITE)]:
        ws.merge_cells(start_row=leg,start_column=c1,end_row=leg,end_column=c2)
        c=ws.cell(row=leg,column=c1,value=lbl)
        c.fill=PatternFill("solid",fgColor=bg)
        c.font=Font(color=fg,bold=True,name=FM,size=SZ); c.alignment=ctr; c.border=bdr
    ws.row_dimensions[leg].height=16; row_num+=1

    sig=row_num
    for label,c1,c2 in [("작성자",1,3),("검토자",4,7),("승인자",8,11)]:
        ws.merge_cells(start_row=sig,start_column=c1,end_row=sig,end_column=c2)
        c=ws.cell(row=sig,column=c1,value=f"{label}:                              (인)")
        c.border=bdr; c.alignment=Alignment(horizontal="left",vertical="top")
        c.font=f_sig; c.fill=fill_sig
    ws.row_dimensions[sig].height=ROW_H*1.6; row_num+=1

    ws.print_area=f"A1:{get_column_letter(PC+3)}{row_num-1}"
    ws.page_setup.orientation="landscape"; ws.page_setup.paperSize=9
    ws.page_setup.fitToPage=True; ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=0
    ws.page_margins=PageMargins(left=0.4,right=0.4,top=0.5,bottom=0.5,header=0.2,footer=0.2)
    ws.print_title_rows="1:3"; ws.freeze_panes="A4"
    out=io.BytesIO(); wb.save(out); return out.getvalue()
